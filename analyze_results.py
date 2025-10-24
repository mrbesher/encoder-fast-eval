#!/usr/bin/env python3

import argparse
import json
import os
import sys
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Tuple, Any

try:
    from rich.console import Console
    from rich.table import Table
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ResultsAnalyzer:
    def __init__(self, results_dir: str):
        self.results_dir = results_dir
        self.results = self._load_results()
        self.metadata = self._compute_metadata()

    def _parse_model_name(self, model: str) -> str:
        if model.startswith("localhost__"):
            return model.split("/")[-1].replace("localhost__", "")
        return model.replace("__", "/")

    def _load_results(self) -> Dict[str, Dict[str, Any]]:
        results = defaultdict(lambda: defaultdict(dict))

        for model_dir in Path(self.results_dir).iterdir():
            if not model_dir.is_dir():
                continue

            model_name = self._parse_model_name(model_dir.name)

            for config_dir in model_dir.iterdir():
                if not config_dir.is_dir():
                    continue

                results_file = config_dir / "results.jsonl"
                if not results_file.exists():
                    continue

                with open(results_file) as f:
                    for line in f:
                        if not line.strip():
                            continue

                        data = json.loads(line)
                        if "runs" not in data:
                            continue

                        dataset_key = data["dataset"]
                        if data.get("subset"):
                            dataset_key += f" ({data['subset'][:6]})"

                        results[model_name][dataset_key] = {
                            "type": data["type"],
                            "metrics": data["metrics"],
                            "main_score": data["metrics"].get("main_score_mean", 0.0),
                            "main_score_std": data["metrics"].get("main_score_std", 0.0),
                            "num_labels": data["num_labels"]
                        }

        return results

    def _compute_metadata(self) -> Dict[str, Any]:
        all_datasets = sorted(set(d for model_data in self.results.values() for d in model_data))
        all_task_types = sorted(set(d["type"] for model_data in self.results.values() for d in model_data.values()))

        dataset_tasks = {}
        for dataset in all_datasets:
            for model_data in self.results.values():
                if dataset in model_data:
                    dataset_tasks[dataset] = model_data[dataset]["type"]
                    break

        task_averages = defaultdict(lambda: defaultdict(list))
        for model, datasets in self.results.items():
            for dataset_name, data in datasets.items():
                task_averages[model][data["type"]].append(data["main_score"])

        for model in task_averages:
            for task_type in task_averages[model]:
                scores = task_averages[model][task_type]
                task_averages[model][task_type] = sum(scores) / len(scores)

        best_performers = {}
        for model, datasets in self.results.items():
            for dataset_name, data in datasets.items():
                key = (data["type"], dataset_name)
                score = data["main_score"]
                if key not in best_performers or score > best_performers[key][1]:
                    best_performers[key] = (model, score)
        best_performers = {k: v[0] for k, v in best_performers.items()}

        best_task_averages = {}
        for task_type in all_task_types:
            best_score = -1
            best_model = None
            for model in self.results:
                if task_type in task_averages[model]:
                    score = task_averages[model][task_type]
                    if score > best_score:
                        best_score = score
                        best_model = model
            if best_model:
                best_task_averages[(task_type,)] = best_model

        return {
            "all_datasets": all_datasets,
            "all_task_types": all_task_types,
            "dataset_tasks": dataset_tasks,
            "task_averages": dict(task_averages),
            "best_performers": best_performers,
            "best_task_averages": best_task_averages
        }

    def _format_score(self, score: float, std: float = 0.0, show_std: bool = True) -> str:
        if show_std and std > 0:
            return f"{score:.3f} ± {std:.3f}"
        return f"{score:.3f}"

    def _truncate_name(self, name: str, max_len: int = 12) -> str:
        return name[:max_len] + "..." if len(name) > max_len else name

    def print_console(self, full_table: bool = False, no_bold: bool = False):
        if not RICH_AVAILABLE:
            print("Rich library not installed. Run: uv add rich")
            return

        console = Console()
        table = Table(title="Task Averages" if not full_table else "Evaluation Results")
        table.add_column("Model", style="cyan", no_wrap=True)

        for task_type in self.metadata["all_task_types"]:
            table.add_column(f"{task_type}{' (avg)' if full_table else ''}", style="magenta", justify="center")

        if full_table:
            for dataset in self.metadata["all_datasets"]:
                table.add_column(self._truncate_name(dataset, 12), justify="center")

        for model in sorted(self.results.keys()):
            row = [model]

            for task_type in self.metadata["all_task_types"]:
                if task_type in self.metadata["task_averages"][model]:
                    avg_score = self.metadata["task_averages"][model][task_type]
                    is_best = self.metadata["best_task_averages"].get((task_type,)) == model
                    score_str = f"{avg_score:.3f}"
                    if is_best and not no_bold:
                        score_str = f"[bold yellow]{score_str}[/bold yellow]"
                    row.append(score_str)
                else:
                    row.append("-")

            if full_table:
                for dataset in self.metadata["all_datasets"]:
                    if dataset in self.results[model]:
                        data = self.results[model][dataset]
                        score_str = self._format_score(data["main_score"], data.get("main_score_std", 0.0))
                        is_best = self.metadata["best_performers"].get((data["type"], dataset)) == model
                        if is_best and not no_bold:
                            score_str = f"[bold yellow]{score_str}[/bold yellow]"
                        row.append(score_str)
                    else:
                        row.append("-")

            table.add_row(*row)

        console.print(table)

    def print_markdown(self, no_bold: bool = False):
        rows = []
        header = ["Model"]

        for task_type in self.metadata["all_task_types"]:
            header.append(f"{task_type.title()} Avg")

        header.extend(self.metadata["all_datasets"])
        rows.append("| " + " | ".join(header) + " |")
        rows.append("|" + "---|" * len(header))

        task_row = ["-"]
        for _ in self.metadata["all_task_types"]:
            task_row.append("-")
        for dataset in self.metadata["all_datasets"]:
            task_row.append(self.metadata["dataset_tasks"].get(dataset, "-"))
        rows.append("| " + " | ".join(task_row) + " |")

        for model in sorted(self.results.keys()):
            row = [model]

            for task_type in self.metadata["all_task_types"]:
                if task_type in self.metadata["task_averages"][model]:
                    avg_score = self.metadata["task_averages"][model][task_type]
                    is_best = self.metadata["best_task_averages"].get((task_type,)) == model
                    score_str = f"{avg_score:.3f}"
                    if is_best and not no_bold:
                        score_str = f"**{score_str}**"
                    row.append(score_str)
                else:
                    row.append("-")

            for dataset in self.metadata["all_datasets"]:
                if dataset in self.results[model]:
                    data = self.results[model][dataset]
                    score_str = self._format_score(data["main_score"], data.get("main_score_std", 0.0))
                    is_best = self.metadata["best_performers"].get((data["type"], dataset)) == model
                    if is_best and not no_bold:
                        score_str = f"**{score_str}**"
                    row.append(score_str)
                else:
                    row.append("-")

            rows.append("| " + " | ".join(row) + " |")

        print("\n".join(rows))

    def print_csv(self):
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        header = ["Model"]
        for task_type in self.metadata["all_task_types"]:
            header.append(f"{task_type.title()} Avg")
        header.extend(self.metadata["all_datasets"])
        writer.writerow(header)

        task_row = ["-"]
        for _ in self.metadata["all_task_types"]:
            task_row.append("-")
        for dataset in self.metadata["all_datasets"]:
            task_row.append(self.metadata["dataset_tasks"].get(dataset, "-"))
        writer.writerow(task_row)

        for model in sorted(self.results.keys()):
            row = [model]

            for task_type in self.metadata["all_task_types"]:
                if task_type in self.metadata["task_averages"][model]:
                    avg_score = self.metadata["task_averages"][model][task_type]
                    row.append(f"{avg_score:.3f}")
                else:
                    row.append("")

            for dataset in self.metadata["all_datasets"]:
                if dataset in self.results[model]:
                    data = self.results[model][dataset]
                    score = data["main_score"]
                    std = data.get("main_score_std", 0.0)
                    row.append(f"{score:.4f}")
                    row.append(f"{std:.4f}" if std > 0 else "")
                else:
                    row.append("")
                    row.append("")

            writer.writerow(row)

        print(output.getvalue().strip())

    def print_json(self):
        output = {
            "_dataset_task_types": self.metadata["dataset_tasks"],
            "task_averages": {}
        }

        for model in sorted(self.results.keys()):
            model_data = {"datasets": {}, "task_averages": {}}

            for dataset, data in self.results[model].items():
                task_type = data["type"]
                is_best = self.metadata["best_performers"].get((task_type, dataset)) == model

                model_data["datasets"][dataset] = {
                    "type": task_type,
                    "main_score": data["main_score"],
                    "main_score_std": data.get("main_score_std", 0.0),
                    "metrics": data["metrics"],
                    "is_best": is_best
                }

            for task_type, avg_score in self.metadata["task_averages"].get(model, {}).items():
                model_data["task_averages"][task_type] = {
                    "score": avg_score,
                    "datasets": [d for d, data in self.results[model].items() if data["type"] == task_type]
                }

            output[model] = model_data

        print(json.dumps(output, indent=2))

    def print_excel(self, output_file: str = "results.xlsx"):
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl library not installed. Run: uv add openpyxl", file=sys.stderr)
            sys.exit(1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        styles = {
            "header_font": Font(bold=True),
            "header_fill": PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
            "task_type_font": Font(italic=True, color="666666"),
            "center_align": Alignment(horizontal='center', vertical='center'),
            "best_font": Font(bold=True),
            "gray_font": Font(color="888888", size=9),
            "value_fill": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
            "thin_border": Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        }

        col = 1
        ws.cell(row=1, column=col, value="Model").font = styles["header_font"]
        ws.cell(row=1, column=col).fill = styles["header_fill"]
        ws.cell(row=1, column=col).alignment = styles["center_align"]
        col += 1

        task_avg_columns = {}
        task_avg_std_columns = {}
        for task_type in self.metadata["all_task_types"]:
            task_avg_columns[task_type] = col
            task_avg_std_columns[task_type] = col + 1

            ws.cell(row=1, column=col, value=f"{task_type.title()}").font = styles["header_font"]
            ws.cell(row=1, column=col).fill = styles["header_fill"]
            ws.cell(row=1, column=col).alignment = styles["center_align"]
            ws.cell(row=1, column=col+1).font = styles["header_font"]
            ws.cell(row=1, column=col+1).fill = styles["header_fill"]
            ws.cell(row=1, column=col+1).alignment = styles["center_align"]

            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            col += 2

        dataset_columns = {}
        dataset_std_columns = {}
        for dataset in self.metadata["all_datasets"]:
            dataset_columns[dataset] = col
            dataset_std_columns[dataset] = col + 1

            ws.cell(row=1, column=col, value=dataset).font = styles["header_font"]
            ws.cell(row=1, column=col).fill = styles["header_fill"]
            ws.cell(row=1, column=col).alignment = styles["center_align"]

            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            col += 2

        ws.cell(row=2, column=1, value="-").font = styles["task_type_font"]
        ws.cell(row=2, column=1).alignment = styles["center_align"]

        for task_type in self.metadata["all_task_types"]:
            ws.cell(row=2, column=task_avg_columns[task_type], value="avg").font = styles["task_type_font"]
            ws.cell(row=2, column=task_avg_columns[task_type]).alignment = styles["center_align"]
            ws.cell(row=2, column=task_avg_std_columns[task_type], value="±").font = styles["task_type_font"]
            ws.cell(row=2, column=task_avg_std_columns[task_type]).alignment = styles["center_align"]

        for dataset in self.metadata["all_datasets"]:
            ws.cell(row=2, column=dataset_columns[dataset], value="avg").font = styles["task_type_font"]
            ws.cell(row=2, column=dataset_columns[dataset]).alignment = styles["center_align"]
            ws.cell(row=2, column=dataset_std_columns[dataset], value="±").font = styles["task_type_font"]
            ws.cell(row=2, column=dataset_std_columns[dataset]).alignment = styles["center_align"]

        ws.cell(row=3, column=1, value="Task").font = styles["task_type_font"]
        ws.cell(row=3, column=1).alignment = styles["center_align"]

        for task_type in self.metadata["all_task_types"]:
            ws.cell(row=3, column=task_avg_columns[task_type], value="-").font = styles["task_type_font"]
            ws.cell(row=3, column=task_avg_columns[task_type]).alignment = styles["center_align"]
            ws.cell(row=3, column=task_avg_std_columns[task_type], value="-").font = styles["task_type_font"]
            ws.cell(row=3, column=task_avg_std_columns[task_type]).alignment = styles["center_align"]

        for dataset in self.metadata["all_datasets"]:
            task_type = self.metadata["dataset_tasks"].get(dataset, "-")
            ws.cell(row=3, column=dataset_columns[dataset], value=task_type).font = styles["task_type_font"]
            ws.cell(row=3, column=dataset_columns[dataset]).alignment = styles["center_align"]
            ws.cell(row=3, column=dataset_std_columns[dataset], value=task_type).font = styles["task_type_font"]
            ws.cell(row=3, column=dataset_std_columns[dataset]).alignment = styles["center_align"]

        data_start_row = 4
        for model_idx, model in enumerate(sorted(self.results.keys())):
            row = data_start_row + model_idx
            ws.cell(row=row, column=1, value=model)

            for task_type in self.metadata["all_task_types"]:
                ws.cell(row=row, column=task_avg_columns[task_type])
                ws.cell(row=row, column=task_avg_std_columns[task_type])

            for dataset in self.metadata["all_datasets"]:
                col = dataset_columns[dataset]
                std_col = dataset_std_columns[dataset]

                if dataset in self.results[model]:
                    data = self.results[model][dataset]
                    score = data["main_score"]
                    std = data.get("main_score_std", 0.0)
                    is_best = self.metadata["best_performers"].get((data["type"], dataset)) == model

                    score_cell = ws.cell(row=row, column=col, value=score)
                    if is_best:
                        score_cell.font = styles["best_font"]
                    score_cell.alignment = styles["center_align"]
                    score_cell.number_format = "0.000"

                    if std >= 0:
                        std_cell = ws.cell(row=row, column=std_col, value=std)
                        std_cell.font = styles["gray_font"]
                        std_cell.alignment = styles["center_align"]
                        std_cell.number_format = "0.000"
                        std_cell.fill = styles["value_fill"]

        for model_idx, model in enumerate(sorted(self.results.keys())):
            row = data_start_row + model_idx

            for task_type in self.metadata["all_task_types"]:
                avg_col = task_avg_columns[task_type]
                std_col = task_avg_std_columns[task_type]

                cols_for_task = []
                std_cols_for_task = []
                for dataset, task in self.metadata["dataset_tasks"].items():
                    if task == task_type:
                        cols_for_task.append(dataset_columns[dataset])
                        std_cols_for_task.append(dataset_std_columns[dataset])

                if cols_for_task:
                    if len(cols_for_task) == 1:
                        formula = f"=IF(ISNUMBER({get_column_letter(cols_for_task[0])}{row}), {get_column_letter(cols_for_task[0])}{row}, \"\")"
                    else:
                        col_refs = [f"{get_column_letter(c)}{row}" for c in cols_for_task]
                        formula = f"=IFERROR(AVERAGE({','.join(col_refs)}), \"\")"

                    ws.cell(row=row, column=avg_col, value=formula)
                    ws.cell(row=row, column=avg_col).number_format = "0.000"
                    ws.cell(row=row, column=avg_col).alignment = styles["center_align"]

                    if len(std_cols_for_task) > 0:
                        std_refs = [f"{get_column_letter(c)}{row}" for c in std_cols_for_task]
                        if len(std_refs) == 1:
                            std_formula = f"=IF(ISNUMBER({std_refs[0]}), {std_refs[0]}, \"\")"
                        else:
                            std_formula = f"=IFERROR(AVERAGE({','.join(std_refs)}), \"\")"

                        ws.cell(row=row, column=std_col, value=std_formula)
                        ws.cell(row=row, column=std_col).number_format = "0.000"
                        ws.cell(row=row, column=std_col).alignment = styles["center_align"]
                        ws.cell(row=row, column=std_col).font = styles["gray_font"]
                        ws.cell(row=row, column=std_col).fill = styles["value_fill"]

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20 if col == 1 else 10

        for row in range(1, 4):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).border = styles["thin_border"]

        ws.freeze_panes = "A4"
        wb.save(output_file)
        print(f"Excel file saved to: {output_file}")
        print(f"Note: Average columns use Excel formulas and update automatically when values change.")


def main():
    parser = argparse.ArgumentParser(
        description="Analyze encoder evaluation results",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s                          # Show task averages in console
  %(prog)s console --full-table     # Show full table with all datasets
  %(prog)s excel -o results.xlsx    # Export to Excel
  %(prog)s markdown > results.md    # Export to Markdown
  %(prog)s csv > results.csv        # Export to CSV
  %(prog)s json > results.json      # Export to JSON
        """
    )

    parser.add_argument("format", choices=["console", "markdown", "csv", "json", "excel"],
                       nargs="?", default="console", help="Output format (default: console)")

    parser.add_argument("--results-dir", default="results",
                       help="Results directory path (default: results)")

    parser.add_argument("--full-table", action="store_true",
                       help="Show full table with all datasets (console format only)")

    parser.add_argument("--no-bold", action="store_true",
                       help="Don't highlight best performers")

    parser.add_argument("-o", "--output", type=str,
                       help="Output file name (Excel format only)")

    args = parser.parse_args()

    if not os.path.exists(args.results_dir):
        print(f"Error: Results directory '{args.results_dir}' not found", file=sys.stderr)
        sys.exit(1)

    analyzer = ResultsAnalyzer(args.results_dir)

    if not analyzer.results:
        print("No results found", file=sys.stderr)
        sys.exit(1)

    if args.format == "console":
        analyzer.print_console(full_table=args.full_table, no_bold=args.no_bold)
    elif args.format == "markdown":
        analyzer.print_markdown(no_bold=args.no_bold)
    elif args.format == "csv":
        analyzer.print_csv()
    elif args.format == "json":
        analyzer.print_json()
    elif args.format == "excel":
        analyzer.print_excel(output_file=args.output or "results.xlsx")


if __name__ == "__main__":
    main()